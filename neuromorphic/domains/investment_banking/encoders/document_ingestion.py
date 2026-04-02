"""
DocumentIngestion — Extract structured financial data from documents
====================================================================
Converts PDFs, Excel files, and plain text into FinancialChunk objects
that the FinancialEncoder can encode into spike trains.

Supported formats:
  - PDF (via pdfplumber if installed, else text-based fallback)
  - Excel .xlsx (via openpyxl)
  - Plain text / markdown
  - Web page HTML (stripped)
  - YouTube transcripts (plain text)
"""

from __future__ import annotations

import re
import os
from dataclasses import dataclass, field
from typing import Iterator

import numpy as np

from neuromorphic.domains.investment_banking.encoders.financial_encoder import FinancialChunk


# ---------------------------------------------------------------------------
# Financial entity extraction — regex + dictionary (no ML dependency)
# ---------------------------------------------------------------------------

# Regex patterns for common financial quantities
_CURRENCY_PATTERNS = [
    (r'\$\s*([\d,]+\.?\d*)\s*([BMK]?)', "usd"),
    (r'([\d,]+\.?\d*)\s*(?:million|MM|mn)\s*(?:USD|dollars?)?', "usd_million"),
    (r'([\d,]+\.?\d*)\s*(?:billion|bn)\s*(?:USD|dollars?)?', "usd_billion"),
]
_PCT_PATTERN    = re.compile(r'([\d]+\.?\d*)\s*%')
_MULTIPLE_PATTERN = re.compile(r'([\d]+\.?\d*)\s*x\b', re.IGNORECASE)
_YEAR_PATTERN   = re.compile(r'\b(20[12]\d)\b')

# Canonical financial metric names
_METRIC_ALIASES: dict[str, str] = {
    "revenue": "revenue", "sales": "revenue", "turnover": "revenue",
    "ebitda": "ebitda", "operating profit": "ebit", "ebit": "ebit",
    "net income": "net_income", "net profit": "net_income",
    "free cash flow": "fcf", "fcf": "fcf", "unlevered fcf": "ufcf",
    "capex": "capex", "capital expenditure": "capex",
    "total debt": "total_debt", "net debt": "net_debt",
    "cash": "cash", "enterprise value": "ev", "equity value": "equity_value",
    "market cap": "market_cap", "market capitalization": "market_cap",
    "working capital": "nwc", "nwc": "nwc",
    "interest expense": "interest_expense", "tax rate": "tax_rate",
    "depreciation": "depreciation", "amortization": "amortization",
    "gross profit": "gross_profit", "gross margin": "gross_margin",
    "ebitda margin": "ebitda_margin", "operating margin": "ebitda_margin",
    "revenue growth": "revenue_growth", "growth rate": "revenue_growth",
    "discount rate": "discount_rate", "wacc": "wacc", "irr": "irr",
    "ev/ebitda": "ev_ebitda", "ev/revenue": "ev_revenue", "p/e": "pe_ratio",
    "entry multiple": "entry_multiple", "exit multiple": "exit_multiple",
    "leverage": "leverage_ratio", "debt/ebitda": "leverage_ratio",
    "premium": "premium_pct", "acquisition premium": "premium_pct",
    "synergies": "synergy_pct",
}


class FinancialEntityExtractor:
    """
    Pattern-based NER for financial entities.
    Extracts concepts (IB terms) and numerical values from text.
    """

    def __init__(self, vocabulary: dict[str, int]):
        self.vocab = vocabulary

    def extract_entities(self, text: str) -> tuple[list[str], dict[str, float]]:
        """
        Extract (concepts, numerical_values) from text.

        Returns
        -------
        concepts : list[str]  — IB vocabulary terms found
        values   : dict[str, float]  — named numerical metrics
        """
        concepts = self._extract_concepts(text)
        values   = self._extract_numbers(text)
        return concepts, values

    def _extract_concepts(self, text: str) -> list[str]:
        text_lower = text.lower()
        found = []
        for term in self.vocab:
            if term.replace("_", " ") in text_lower or term in text_lower:
                found.append(term)
        return list(dict.fromkeys(found))   # preserve order, deduplicate

    def _extract_numbers(self, text: str) -> dict[str, float]:
        values: dict[str, float] = {}

        # Currency amounts
        for pattern_str, currency in _CURRENCY_PATTERNS:
            for m in re.finditer(pattern_str, text, re.IGNORECASE):
                amount_str = m.group(1).replace(",", "")
                try:
                    amount = float(amount_str)
                    suffix = m.group(2).upper() if len(m.groups()) > 1 else ""
                    if suffix == "B":   amount *= 1e9
                    elif suffix == "M": amount *= 1e6
                    elif suffix == "K": amount *= 1e3
                    elif "billion" in pattern_str: amount *= 1e9
                    elif "million" in pattern_str: amount *= 1e6
                    # Try to name it from context
                    ctx = text[max(0, m.start()-30):m.start()].lower()
                    name = self._name_from_context(ctx)
                    if name:
                        values[name] = amount
                except ValueError:
                    pass

        # Percentages
        for m in _PCT_PATTERN.finditer(text):
            pct   = float(m.group(1)) / 100.0
            ctx   = text[max(0, m.start()-40):m.start()].lower()
            name  = self._name_from_context(ctx)
            if name:
                values[name] = pct

        # Multiples (Nx)
        for m in _MULTIPLE_PATTERN.finditer(text):
            val  = float(m.group(1))
            ctx  = text[max(0, m.start()-40):m.start()].lower()
            name = self._name_from_context(ctx)
            if name:
                values[name] = val

        return values

    def _name_from_context(self, context: str) -> str | None:
        """Guess metric name from preceding text context."""
        for alias, canonical in _METRIC_ALIASES.items():
            if alias in context:
                return canonical
        return None

    def extract_relationships(self, text: str) -> list[tuple[str, str, str]]:
        """Extract (entity, relation, entity) triples."""
        triples = []
        # Simple pattern: "X is Y", "X has Y", "X acquires Y"
        relation_patterns = [
            (r'(\w+(?:\s+\w+)?)\s+acquires?\s+(\w+(?:\s+\w+)?)', "acquires"),
            (r'(\w+(?:\s+\w+)?)\s+(?:has|had)\s+an?\s+(\w+)', "has"),
            (r'(\w+(?:\s+\w+)?)\s+(?:is|was)\s+(?:a|an)?\s*(\w+(?:\s+\w+)?)', "is"),
        ]
        for pattern, relation in relation_patterns:
            for m in re.finditer(pattern, text, re.IGNORECASE):
                triples.append((m.group(1).strip(), relation, m.group(2).strip()))
        return triples[:20]  # cap at 20 per chunk


# ---------------------------------------------------------------------------
# Document ingestion pipeline
# ---------------------------------------------------------------------------

class DocumentIngestion:
    """
    Extracts FinancialChunks from raw documents.

    Parameters
    ----------
    config : ib_config module
    """

    CHUNK_SIZE_CHARS = 1500   # ~500 tokens per chunk

    def __init__(self, config):
        self.cfg       = config
        self.extractor = FinancialEntityExtractor(config.IB_VOCABULARY)

    def extract(self, file_path: str, doc_type: str = "auto") -> list[FinancialChunk]:
        """
        Main entry point. Detects format and extracts chunks.
        """
        if doc_type == "auto":
            ext = os.path.splitext(file_path)[1].lower()
            if ext == ".pdf":
                doc_type = "pdf"
            elif ext in (".xlsx", ".xls", ".xlsm"):
                doc_type = "excel"
            elif ext in (".txt", ".md", ".rst"):
                doc_type = "text"
            elif ext in (".html", ".htm"):
                doc_type = "html"
            else:
                doc_type = "text"

        if doc_type == "pdf":
            return self._extract_pdf(file_path)
        elif doc_type == "excel":
            return self._extract_excel(file_path)
        elif doc_type == "html":
            return self._extract_html(file_path)
        else:
            return self._extract_text(file_path)

    def extract_from_string(
        self,
        text: str,
        source: str = "string",
        chunk_type: str = "text",
    ) -> list[FinancialChunk]:
        """Parse raw text directly (for web content, transcripts, etc.)."""
        return list(self._chunk_text(text, source, chunk_type))

    # ------------------------------------------------------------------
    # Format-specific extractors
    # ------------------------------------------------------------------

    def _extract_pdf(self, path: str) -> list[FinancialChunk]:
        try:
            import pdfplumber
            chunks = []
            with pdfplumber.open(path) as pdf:
                for page_num, page in enumerate(pdf.pages):
                    text = page.extract_text() or ""
                    tables = page.extract_tables() or []
                    # Text chunks
                    for chunk in self._chunk_text(text, f"{path}:p{page_num}", "text"):
                        chunks.append(chunk)
                    # Table chunks
                    for table in tables:
                        chunk = self._table_to_chunk(table, f"{path}:p{page_num}:table")
                        if chunk:
                            chunks.append(chunk)
            return chunks
        except ImportError:
            # Fallback: treat as text
            return self._extract_text(path)
        except Exception:
            return []

    def _extract_excel(self, path: str) -> list[FinancialChunk]:
        try:
            import openpyxl
            wb     = openpyxl.load_workbook(path, data_only=True)
            chunks = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                # Extract all cell values as a flat text block
                rows = []
                for row in ws.iter_rows(values_only=True):
                    row_text = "  ".join(
                        str(cell) for cell in row if cell is not None
                    )
                    if row_text.strip():
                        rows.append(row_text)
                text = "\n".join(rows)
                # Also build a numeric table
                num_rows = []
                for row in ws.iter_rows(values_only=True):
                    num_row = [
                        float(cell) for cell in row
                        if isinstance(cell, (int, float)) and cell is not None
                    ]
                    if num_row:
                        num_rows.append(num_row)
                table_arr = None
                if num_rows:
                    max_len = max(len(r) for r in num_rows)
                    padded  = [r + [0.0] * (max_len - len(r)) for r in num_rows]
                    table_arr = np.array(padded, dtype=np.float32)
                # Chunk the text
                for chunk in self._chunk_text(text, f"{path}:{sheet_name}", "model"):
                    if table_arr is not None:
                        chunk.table_data = table_arr
                    chunks.append(chunk)
            return chunks
        except ImportError:
            return []
        except Exception:
            return []

    def _extract_html(self, path: str) -> list[FinancialChunk]:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                html = f.read()
            text = re.sub(r"<[^>]+>", " ", html)
            text = re.sub(r"\s+", " ", text).strip()
            return list(self._chunk_text(text, path, "web"))
        except Exception:
            return []

    def _extract_text(self, path: str) -> list[FinancialChunk]:
        try:
            with open(path, "r", encoding="utf-8", errors="ignore") as f:
                text = f.read()
            return list(self._chunk_text(text, path, "text"))
        except Exception:
            return []

    # ------------------------------------------------------------------
    # Chunking helpers
    # ------------------------------------------------------------------

    def _chunk_text(
        self,
        text: str,
        source: str,
        chunk_type: str,
    ) -> Iterator[FinancialChunk]:
        """Split text into chunks and extract entities from each."""
        text = text.strip()
        if not text:
            return
        start = 0
        while start < len(text):
            end        = min(start + self.CHUNK_SIZE_CHARS, len(text))
            chunk_text = text[start:end]
            start      = end

            concepts, values = self.extractor.extract_entities(chunk_text)
            rels             = self.extractor.extract_relationships(chunk_text)

            if concepts or values:
                yield FinancialChunk(
                    concepts          = concepts,
                    numerical_values  = values,
                    relationships     = rels,
                    table_data        = None,
                    source            = source,
                    chunk_type        = chunk_type,
                )

    def _table_to_chunk(self, table: list[list], source: str) -> FinancialChunk | None:
        """Convert an extracted table to a FinancialChunk."""
        flat_text = " ".join(
            str(cell) for row in table for cell in row if cell is not None
        )
        concepts, values = self.extractor.extract_entities(flat_text)
        if not (concepts or values):
            return None
        return FinancialChunk(
            concepts         = concepts,
            numerical_values = values,
            relationships    = [],
            table_data       = None,
            source           = source,
            chunk_type       = "table",
        )
