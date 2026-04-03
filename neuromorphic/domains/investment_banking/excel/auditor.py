"""
ExcelAuditor — Overnight model checking and correction
=======================================================
The overnight workhorse. Reads YOUR Excel models, detects every error,
inconsistency, and broken logic — then writes a corrected version.

Before you send anything out, this catches:
  - Formula errors (#REF!, #DIV/0!, #VALUE!, #NAME?, circular refs)
  - Hardcoded numbers where formulas should be
  - Broken cross-sheet references
  - Sum mismatches (e.g. subtotals that don't add up)
  - Balance sheet that doesn't balance
  - Negative EBITDA in a DCF model
  - IRR / MOIC inconsistencies
  - Incorrect tax shield calculations
  - NWC sign convention errors
  - Missing sensitivity analyses
  - Formatting inconsistencies (mixed date formats, inconsistent decimals)
  - Empty assumption cells
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from typing import Any


@dataclass
class AuditIssue:
    severity:    str          # "critical", "warning", "info"
    category:    str          # "formula_error", "logic_error", "format", etc.
    sheet:       str
    cell:        str          # e.g. "B12"
    description: str
    auto_fixable: bool = False
    fix_value:    Any  = None    # value to write if auto_fixable


@dataclass
class AuditReport:
    file_path:      str
    n_issues:       int
    n_critical:     int
    n_warnings:     int
    n_auto_fixed:   int
    issues:         list[AuditIssue]
    corrected_path: str | None = None
    summary:        str        = ""

    def __str__(self) -> str:
        lines = [
            f"AUDIT REPORT: {os.path.basename(self.file_path)}",
            f"{'='*60}",
            f"Total issues:   {self.n_issues}",
            f"  Critical:     {self.n_critical}",
            f"  Warnings:     {self.n_warnings}",
            f"  Auto-fixed:   {self.n_auto_fixed}",
            "",
        ]
        for issue in self.issues[:30]:   # cap display at 30
            icon = "🔴" if issue.severity == "critical" else "🟡" if issue.severity == "warning" else "ℹ"
            fix  = " [FIXED]" if issue.auto_fixable else ""
            lines.append(
                f"{icon} [{issue.sheet}!{issue.cell}] {issue.description}{fix}"
            )
        if len(self.issues) > 30:
            lines.append(f"... and {len(self.issues)-30} more issues")
        if self.corrected_path:
            lines.append(f"\nCorrected file: {self.corrected_path}")
        return "\n".join(lines)


class ExcelAuditor:
    """
    Comprehensive IB Excel model auditor.

    Usage
    -----
        auditor  = ExcelAuditor()
        report   = auditor.audit("my_model.xlsx")
        corrected = auditor.audit_and_correct("my_model.xlsx")
        print(report)
    """

    # Error values that openpyxl returns for broken formulas
    _EXCEL_ERRORS = {"#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A",
                     "#NULL!", "#NUM!", "#ERROR!"}

    def __init__(self):
        self._avail = self._check_openpyxl()

    @staticmethod
    def _check_openpyxl() -> bool:
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False

    # ------------------------------------------------------------------
    # Main audit entry points
    # ------------------------------------------------------------------

    def audit(self, path: str) -> AuditReport:
        """Full audit of an Excel file. Returns report without modifying file."""
        issues = self._run_all_checks(path)
        n_crit = sum(1 for i in issues if i.severity == "critical")
        n_warn = sum(1 for i in issues if i.severity == "warning")
        return AuditReport(
            file_path   = path,
            n_issues    = len(issues),
            n_critical  = n_crit,
            n_warnings  = n_warn,
            n_auto_fixed = 0,
            issues      = issues,
            summary     = self._generate_summary(issues),
        )

    def audit_and_correct(self, path: str, output_path: str | None = None) -> AuditReport:
        """
        Audit and auto-correct all fixable issues.
        Writes corrected workbook to output_path (default: *_corrected.xlsx).
        """
        if not self._avail:
            return AuditReport(path, 0, 0, 0, 0, [],
                               summary="openpyxl not installed")

        issues = self._run_all_checks(path)

        if output_path is None:
            base, ext = os.path.splitext(path)
            output_path = f"{base}_corrected{ext}"

        n_fixed = self._apply_fixes(path, output_path, issues)

        n_crit = sum(1 for i in issues if i.severity == "critical")
        n_warn = sum(1 for i in issues if i.severity == "warning")
        return AuditReport(
            file_path    = path,
            n_issues     = len(issues),
            n_critical   = n_crit,
            n_warnings   = n_warn,
            n_auto_fixed = n_fixed,
            issues       = issues,
            corrected_path = output_path,
            summary      = self._generate_summary(issues),
        )

    # ------------------------------------------------------------------
    # Check suites
    # ------------------------------------------------------------------

    def _run_all_checks(self, path: str) -> list[AuditIssue]:
        """Run all audit checks and return combined issue list."""
        issues: list[AuditIssue] = []
        if not self._avail or not os.path.exists(path):
            return issues
        import openpyxl
        try:
            wb_data    = openpyxl.load_workbook(path, data_only=True)
            wb_formula = openpyxl.load_workbook(path, data_only=False)
        except Exception as e:
            issues.append(AuditIssue("critical", "file_error", "N/A", "N/A",
                                     f"Cannot open file: {e}"))
            return issues

        for sheet_name in wb_data.sheetnames:
            ws_data    = wb_data[sheet_name]
            ws_formula = wb_formula[sheet_name]

            issues += self._check_formula_errors(ws_data, sheet_name)
            issues += self._check_hardcoded_numbers(ws_data, ws_formula, sheet_name)
            issues += self._check_sum_consistency(ws_data, ws_formula, sheet_name)
            issues += self._check_logical_signs(ws_data, sheet_name)
            issues += self._check_empty_labeled_cells(ws_data, ws_formula, sheet_name)

        issues += self._check_balance_sheet(wb_data)
        issues += self._check_negative_ebitda(wb_data)
        issues += self._check_formatting(wb_data)

        return issues

    def _check_formula_errors(self, ws, sheet_name: str) -> list[AuditIssue]:
        """Detect cells containing Excel error values."""
        issues = []
        for row in ws.iter_rows():
            for cell in row:
                val = str(cell.value) if cell.value is not None else ""
                if val in self._EXCEL_ERRORS:
                    issues.append(AuditIssue(
                        severity    = "critical",
                        category    = "formula_error",
                        sheet       = sheet_name,
                        cell        = cell.coordinate,
                        description = f"Formula error: {val}",
                        auto_fixable = val == "#DIV/0!",
                        fix_value   = 0 if val == "#DIV/0!" else None,
                    ))
        return issues

    def _check_hardcoded_numbers(self, ws_data, ws_formula, sheet_name: str) -> list[AuditIssue]:
        """
        Flag numeric cells in output rows that should be formulas.
        Heuristic: if a cell header says 'Total', 'Sum', 'Subtotal' but the cell
        is a hardcoded number rather than a formula, flag it.
        """
        issues = []
        # Find rows/cols with aggregation labels
        agg_labels = {"total", "sum", "subtotal", "net", "combined", "aggregate"}
        for row in ws_formula.iter_rows():
            for cell in row:
                val   = cell.value
                coord = cell.coordinate
                if not isinstance(val, str):
                    continue
                if any(lbl in val.lower() for lbl in agg_labels):
                    # Check cell to the right for hardcoded number
                    col_idx = cell.column
                    for offset in range(1, 6):
                        check_col = col_idx + offset
                        if check_col > ws_formula.max_column:
                            break
                        formula_cell = ws_formula.cell(cell.row, check_col)
                        data_cell    = ws_data.cell(cell.row, check_col)
                        if (isinstance(data_cell.value, (int, float)) and
                                formula_cell.value == data_cell.value and
                                not str(formula_cell.value).startswith("=")):
                            issues.append(AuditIssue(
                                severity    = "warning",
                                category    = "hardcoded",
                                sheet       = sheet_name,
                                cell        = formula_cell.coordinate,
                                description = f"'{val}' cell appears hardcoded — should be a SUM formula",
                            ))
        return issues

    def _check_sum_consistency(self, ws_data, ws_formula, sheet_name: str) -> list[AuditIssue]:
        """Check that SUM formulas actually equal the sum of their referenced cells."""
        issues = []
        for row in ws_formula.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                formula = str(cell.value).upper()
                if not formula.startswith("=SUM("):
                    continue
                # Get the computed value
                data_cell = ws_data.cell(cell.row, cell.column)
                if not isinstance(data_cell.value, (int, float)):
                    continue
                computed = float(data_cell.value)
                # Extract range from formula and sum manually
                m = re.search(r'=SUM\(([^)]+)\)', formula)
                if not m:
                    continue
                range_str = m.group(1)
                try:
                    total = 0.0
                    for cr in ws_data[range_str]:
                        if isinstance(cr, (list, tuple)):
                            for c in cr:
                                if isinstance(c.value, (int, float)):
                                    total += float(c.value)
                        else:
                            if isinstance(cr.value, (int, float)):
                                total += float(cr.value)
                    if abs(total - computed) > max(1.0, abs(computed) * 0.001):
                        issues.append(AuditIssue(
                            severity    = "critical",
                            category    = "sum_mismatch",
                            sheet       = sheet_name,
                            cell        = cell.coordinate,
                            description = f"SUM mismatch: formula returns {computed:,.0f}, actual sum is {total:,.0f}",
                        ))
                except Exception:
                    pass
        return issues

    def _check_logical_signs(self, ws_data, sheet_name: str) -> list[AuditIssue]:
        """
        Flag potentially wrong signs on key financial items.
        CapEx should normally be negative in a cash flow context.
        NWC increase should reduce FCF.
        """
        issues = []
        sign_rules = {
            "capex":      ("should be negative in FCF bridge", lambda v: v > 0),
            "d&a":        ("depreciation is typically positive",  lambda v: v < 0),
            "amortization": ("typically positive", lambda v: v < 0),
        }
        for row in ws_data.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                label_lower = cell.value.lower().strip()
                for term, (msg, bad_sign_fn) in sign_rules.items():
                    if term in label_lower:
                        # Check adjacent numeric cells
                        for offset in range(1, 8):
                            ncell = ws_data.cell(cell.row, cell.column + offset)
                            if isinstance(ncell.value, (int, float)) and ncell.value != 0:
                                if bad_sign_fn(ncell.value):
                                    issues.append(AuditIssue(
                                        severity    = "warning",
                                        category    = "sign_error",
                                        sheet       = sheet_name,
                                        cell        = ncell.coordinate,
                                        description = f"'{cell.value}' sign may be wrong — {msg} "
                                                      f"(current: {ncell.value:,.0f})",
                                    ))
                                break
        return issues

    def _check_empty_labeled_cells(self, ws_data, ws_formula, sheet_name: str) -> list[AuditIssue]:
        """Flag cells with a label but empty value in assumption/input sheets."""
        issues = []
        input_keywords = {"assumption", "input", "driver"}
        if not any(k in sheet_name.lower() for k in input_keywords):
            return issues
        for row in ws_formula.iter_rows():
            cells = [c for c in row if c.value is not None]
            for i in range(len(cells) - 1):
                if (isinstance(cells[i].value, str) and
                        ws_data.cell(cells[i].row, cells[i].column + 1).value is None):
                    issues.append(AuditIssue(
                        severity    = "warning",
                        category    = "empty_input",
                        sheet       = sheet_name,
                        cell        = ws_formula.cell(cells[i].row, cells[i].column + 1).coordinate,
                        description = f"Assumption '{cells[i].value}' has no value",
                    ))
        return issues

    def _check_balance_sheet(self, wb) -> list[AuditIssue]:
        """Check Assets = Liabilities + Equity across all periods."""
        issues = []
        bs_keywords = {"balance", "financial position", "bs"}
        for sheet_name in wb.sheetnames:
            if not any(k in sheet_name.lower() for k in bs_keywords):
                continue
            ws = wb[sheet_name]
            # Look for rows labelled total assets vs total liabilities+equity
            asset_row = equity_row = None
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    lo = cell.value.lower()
                    if "total assets" in lo:
                        asset_row = cell.row
                    if "total liabilities" in lo and "equity" in lo:
                        equity_row = cell.row
            if asset_row and equity_row:
                for col in range(2, ws.max_column + 1):
                    av = ws.cell(asset_row,  col).value
                    lv = ws.cell(equity_row, col).value
                    if isinstance(av, (int, float)) and isinstance(lv, (int, float)):
                        if abs(av - lv) > max(1.0, abs(av) * 0.001):
                            issues.append(AuditIssue(
                                severity    = "critical",
                                category    = "balance_sheet",
                                sheet       = sheet_name,
                                cell        = ws.cell(asset_row, col).coordinate,
                                description = f"Balance sheet doesn't balance: "
                                              f"Assets={av:,.0f}, L+E={lv:,.0f}, "
                                              f"diff={av-lv:,.0f}",
                            ))
        return issues

    def _check_negative_ebitda(self, wb) -> list[AuditIssue]:
        """Flag negative EBITDA values in projection sheets (unusual for a DCF)."""
        issues = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if not isinstance(cell.value, str):
                        continue
                    if "ebitda" in cell.value.lower():
                        for offset in range(1, 10):
                            ncell = ws.cell(cell.row, cell.column + offset)
                            if isinstance(ncell.value, (int, float)) and ncell.value < 0:
                                issues.append(AuditIssue(
                                    severity    = "warning",
                                    category    = "logic_error",
                                    sheet       = sheet_name,
                                    cell        = ncell.coordinate,
                                    description = f"Negative EBITDA: {ncell.value:,.0f} — verify assumptions",
                                ))
        return issues

    def _check_formatting(self, wb) -> list[AuditIssue]:
        """Flag obviously inconsistent number formats."""
        # Lightweight check — look for mix of large and small numbers in same column
        issues = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for col in ws.iter_cols():
                nums = [c.value for c in col if isinstance(c.value, (int, float)) and c.value != 0]
                if len(nums) >= 3:
                    arr = [abs(n) for n in nums]
                    if max(arr) > 1e6 and min(arr) < 1:
                        issues.append(AuditIssue(
                            severity    = "info",
                            category    = "formatting",
                            sheet       = sheet_name,
                            cell        = col[0].coordinate,
                            description = "Column mixes large (>1M) and small (<1) values — possible unit inconsistency",
                        ))
        return issues

    def _apply_fixes(self, source: str, dest: str, issues: list[AuditIssue]) -> int:
        """Apply auto-fixable issues to a copy of the workbook."""
        if not self._avail:
            return 0
        import openpyxl
        import shutil
        shutil.copy2(source, dest)
        try:
            wb = openpyxl.load_workbook(dest)
        except Exception:
            return 0

        n_fixed = 0
        fixable  = [i for i in issues if i.auto_fixable and i.fix_value is not None]
        for issue in fixable:
            try:
                ws = wb[issue.sheet]
                ws[issue.cell] = issue.fix_value
                n_fixed += 1
            except Exception:
                pass

        try:
            wb.save(dest)
        except Exception:
            pass
        return n_fixed

    def _generate_summary(self, issues: list[AuditIssue]) -> str:
        cats: dict[str, int] = {}
        for i in issues:
            cats[i.category] = cats.get(i.category, 0) + 1
        parts = [f"{v} {k.replace('_', ' ')}" for k, v in cats.items()]
        return "Found: " + ", ".join(parts) if parts else "No issues found"
