"""
ExcelReader — Parse financial data from Excel workbooks
========================================================
Reads standard IB model layouts: income statement, balance sheet,
cash flow statement, comp tables, assumptions sheets.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from typing import Any

import numpy as np


@dataclass
class SheetData:
    name:        str
    headers:     list[str]
    rows:        list[dict[str, Any]]     # list of {header: value} rows
    numeric_arr: np.ndarray | None        # shape (n_rows, n_cols) if numeric
    named_cells: dict[str, Any]           # {cell_label: value} from named ranges


class ExcelReader:
    """
    Read financial data from .xlsx files.
    Requires openpyxl (optional dependency — graceful degradation if missing).
    """

    def __init__(self):
        self._avail = self._check_openpyxl()

    @staticmethod
    def _check_openpyxl() -> bool:
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False

    def read_workbook(self, path: str) -> dict[str, SheetData]:
        """Read all sheets from an Excel workbook."""
        if not self._avail:
            return {}
        import openpyxl
        try:
            wb    = openpyxl.load_workbook(path, data_only=True)
            result = {}
            for name in wb.sheetnames:
                ws = wb[name]
                result[name] = self._parse_sheet(ws, name)
            return result
        except Exception as e:
            return {}

    def read_financial_statements(self, path: str) -> dict[str, SheetData]:
        """
        Extract income statement, balance sheet, cash flow by sheet name matching.
        Returns subset matching standard names.
        """
        all_sheets = self.read_workbook(path)
        keywords   = {
            "income":    ["income", "p&l", "pnl", "profit", "revenue"],
            "balance":   ["balance", "bs", "financial position"],
            "cashflow":  ["cash", "cf", "cashflow", "cash flow"],
            "inputs":    ["input", "assumption", "driver"],
            "output":    ["output", "summary", "result"],
        }
        matched: dict[str, SheetData] = {}
        for sheet_name, sheet in all_sheets.items():
            low = sheet_name.lower()
            for category, kws in keywords.items():
                if any(k in low for k in kws):
                    matched[category] = sheet
                    break
        return matched

    def read_named_ranges(self, path: str) -> dict[str, Any]:
        """Extract all defined named ranges as {name: value}."""
        if not self._avail:
            return {}
        import openpyxl
        try:
            wb      = openpyxl.load_workbook(path, data_only=True)
            result  = {}
            for name, defn in wb.defined_names.items():
                try:
                    dests = list(defn.destinations)
                    if dests:
                        sheet_name, cell_ref = dests[0]
                        ws  = wb[sheet_name]
                        val = ws[cell_ref.replace("$", "")].value
                        if val is not None:
                            result[name] = val
                except Exception:
                    pass
            return result
        except Exception:
            return {}

    def extract_assumptions(self, path: str) -> dict[str, float]:
        """
        Best-effort extraction of numeric assumptions from a workbook.
        Looks for label–value pairs (cell with text next to cell with number).
        """
        if not self._avail:
            return {}
        import openpyxl
        result: dict[str, float] = {}
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    cells = [c for c in row if c.value is not None]
                    for i in range(len(cells) - 1):
                        label_cell = cells[i]
                        value_cell = cells[i + 1]
                        if isinstance(label_cell.value, str) and isinstance(value_cell.value, (int, float)):
                            key = re.sub(r"[^a-z0-9_]", "_",
                                         label_cell.value.lower().strip()).strip("_")
                            if key and len(key) < 50:
                                result[key] = float(value_cell.value)
        except Exception:
            pass
        return result

    # ------------------------------------------------------------------
    # Internal
    # ------------------------------------------------------------------

    def _parse_sheet(self, ws, name: str) -> SheetData:
        """Parse a worksheet into SheetData."""
        import openpyxl
        rows_raw = list(ws.iter_rows(values_only=True))
        if not rows_raw:
            return SheetData(name=name, headers=[], rows=[], numeric_arr=None, named_cells={})

        # Treat first row as headers if it contains strings
        first = rows_raw[0]
        has_header = any(isinstance(c, str) for c in first if c is not None)
        if has_header:
            headers  = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(first)]
            data_rows = rows_raw[1:]
        else:
            headers  = [f"Col{i}" for i in range(len(first))]
            data_rows = rows_raw

        rows = [
            {h: row[i] if i < len(row) else None for i, h in enumerate(headers)}
            for row in data_rows
        ]

        # Build numeric array
        num_rows = []
        for row in data_rows:
            nums = [float(c) if isinstance(c, (int, float)) else 0.0 for c in row]
            num_rows.append(nums)
        if num_rows:
            arr = np.array(num_rows, dtype=np.float32)
        else:
            arr = None

        return SheetData(name=name, headers=headers, rows=rows,
                         numeric_arr=arr, named_cells={})
