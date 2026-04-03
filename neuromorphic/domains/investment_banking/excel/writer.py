"""
ExcelWriter — Write financial model outputs to formatted Excel
=============================================================
Produces IB-grade formatted workbooks from model results.
Requires openpyxl.
"""

from __future__ import annotations

import os
from typing import Any

import numpy as np


# ---- Color scheme (Goldman/Morgan Stanley style) ----
BLUE_DARK   = "1F3864"    # Header background
BLUE_MID    = "2E75B6"    # Sub-header
BLUE_LIGHT  = "D6E4F0"    # Alternating row
WHITE       = "FFFFFF"
YELLOW_IN   = "FFFF00"    # Input cells
GREEN_OK    = "92D050"    # Positive values
RED_BAD     = "FF0000"    # Negative / warnings
FONT_HEADER = "Calibri"
FONT_BODY   = "Calibri"


class ExcelWriter:
    """Write model results to IB-formatted Excel workbooks."""

    def __init__(self):
        self._avail = self._check_openpyxl()

    @staticmethod
    def _check_openpyxl() -> bool:
        try:
            import openpyxl  # noqa: F401
            return True
        except ImportError:
            return False

    # ------------------------------------------------------------------
    # Public write methods
    # ------------------------------------------------------------------

    def write_dcf(self, result, path: str, assumptions: dict | None = None):
        """Write a DCF result to Excel with projections and sensitivity."""
        if not self._avail:
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, numbers
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "DCF"

        years = len(result.projected_revenue)
        col_labels = [""] + [f"Year {i+1}" for i in range(years)] + ["Terminal"]

        # --- Header ---
        self._write_header(ws, "Discounted Cash Flow Analysis", len(col_labels))

        # --- Projection table ---
        rows_data = [
            ("Revenue",            result.projected_revenue,        False),
            ("EBITDA",             result.projected_ebitda,         False),
            ("Unlevered FCF",      result.projected_fcf,            False),
            ("PV of FCF",          result.pv_fcfs,                  False),
        ]
        r = 3
        for label, vals, is_pct in rows_data:
            ws.cell(r, 1, label)
            for c, v in enumerate(vals, 2):
                ws.cell(r, c, v).number_format = '#,##0'
            r += 1

        r += 1
        ws.cell(r, 1, "Terminal Value")
        ws.cell(r, 2, result.terminal_value).number_format = '#,##0'
        r += 1
        ws.cell(r, 1, "Enterprise Value")
        ws.cell(r, 2, result.enterprise_value).number_format = '#,##0'
        self._bold(ws.cell(r, 2))
        r += 2

        # --- Sensitivity table ---
        ws.cell(r, 1, "Sensitivity: EV by WACC × Terminal Growth")
        self._bold(ws.cell(r, 1))
        r += 1
        # Column headers = terminal growth rates
        ws.cell(r, 1, "WACC \\ TG")
        for c, tg in enumerate(result.sensitivity_tgs, 2):
            ws.cell(r, c, f"{tg*100:.1f}%")
        r += 1
        for ri, w in enumerate(result.sensitivity_waccs):
            ws.cell(r + ri, 1, f"{w*100:.1f}%")
            for ci, val in enumerate(result.sensitivity_table[ri], 2):
                cell = ws.cell(r + ri, ci, round(val))
                cell.number_format = '#,##0'
                # Highlight diagonal (base case approx)
                if ri == 2 and ci == 4:
                    cell.font = Font(bold=True)

        self._autofit(ws)
        wb.save(path)

    def write_lbo(self, result, path: str):
        """Write LBO result with debt schedule and IRR bridge."""
        if not self._avail:
            return
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LBO"

        self._write_header(ws, "Leveraged Buyout Analysis", 8)
        years = len(result.projected_ebitda)

        r = 3
        headers = [""] + [f"Year {i+1}" for i in range(years)]
        for c, h in enumerate(headers, 1):
            ws.cell(r, c, h)
        self._bold(ws.cell(r, 1))
        r += 1

        for label, vals in [
            ("EBITDA",          result.projected_ebitda),
            ("FCF",             result.projected_fcf),
            ("Debt Outstanding",result.debt_schedule),
            ("Interest Expense",result.interest_expense),
        ]:
            ws.cell(r, 1, label)
            for c, v in enumerate(vals, 2):
                ws.cell(r, c, round(v)).number_format = '#,##0'
            r += 1

        r += 1
        for label, val in [
            ("Entry EV",         result.entry_enterprise_value),
            ("Equity In",        result.equity_contribution),
            ("Exit EV",          result.exit_enterprise_value),
            ("Equity Out",       result.exit_equity_value),
            ("IRR",              result.irr),
            ("MOIC",             result.moic),
        ]:
            ws.cell(r, 1, label)
            if label in ("IRR",):
                ws.cell(r, 2, result.irr).number_format = '0.0%'
            elif label == "MOIC":
                ws.cell(r, 2, f"{result.moic:.2f}x")
            else:
                ws.cell(r, 2, round(val)).number_format = '#,##0'
            r += 1

        # Sensitivity
        r += 1
        ws.cell(r, 1, "IRR Sensitivity: Entry × Exit Multiple")
        self._bold(ws.cell(r, 1))
        r += 1
        ws.cell(r, 1, "Entry \\ Exit")
        for c, xm in enumerate(result.sensitivity_exits, 2):
            ws.cell(r, c, f"{xm:.0f}x")
        r += 1
        for ri, em in enumerate(result.sensitivity_entries):
            ws.cell(r + ri, 1, f"{em:.0f}x")
            for ci, irr_val in enumerate(result.sensitivity_irr[ri], 2):
                ws.cell(r + ri, ci, f"{irr_val*100:.1f}%")

        self._autofit(ws)
        wb.save(path)

    def write_credit(self, result, path: str):
        """Write credit analysis."""
        if not self._avail:
            return
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Credit"
        self._write_header(ws, "Credit & Leverage Analysis", 3)
        r = 3
        metrics = [
            ("Total Leverage (Debt/EBITDA)", f"{result.leverage_total:.1f}x"),
            ("Net Leverage",                 f"{result.leverage_net:.1f}x"),
            ("Interest Coverage",            f"{result.interest_coverage:.1f}x"),
            ("DSCR",                         f"{result.dscr:.2f}x"),
            ("Implied Rating",               result.implied_rating),
            ("Debt Capacity (5x)",           result.debt_capacity_5x),
            ("Debt Capacity (6x)",           result.debt_capacity_6x),
            ("Covenant Check",               "PASS" if result.covenants_ok else "BREACH"),
            ("Covenant Headroom",            f"{result.covenant_headroom_pct:.0f}%"),
            ("Risk Level",                   result.risk_level.upper()),
        ]
        for label, val in metrics:
            ws.cell(r, 1, label)
            ws.cell(r, 2, val)
            r += 1
        self._autofit(ws)
        wb.save(path)

    def write_generic(self, data: dict[str, Any], path: str, title: str = "Analysis"):
        """Write any dict of {label: value} rows to Excel."""
        if not self._avail:
            return
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]
        self._write_header(ws, title, 2)
        r = 3
        for k, v in data.items():
            ws.cell(r, 1, k)
            ws.cell(r, 2, v)
            r += 1
        self._autofit(ws)
        wb.save(path)

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------

    def _write_header(self, ws, title: str, n_cols: int):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        cell = ws.cell(1, 1, title)
        cell.font      = Font(name=FONT_HEADER, bold=True, color=WHITE, size=13)
        cell.fill      = PatternFill("solid", fgColor=BLUE_DARK)
        cell.alignment = Alignment(horizontal="center")
        if n_cols > 1:
            ws.merge_cells(start_row=1, start_column=1,
                           end_row=1,   end_column=n_cols)
        ws.row_dimensions[1].height = 22

    def _bold(self, cell):
        try:
            from openpyxl.styles import Font
            cell.font = Font(bold=True)
        except Exception:
            pass

    def _autofit(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = None
            for cell in col:
                try:
                    col_letter = cell.column_letter
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            if col_letter:
                ws.column_dimensions[col_letter].width = min(max_len + 4, 40)
